import os
from openpyxl import load_workbook


def search_keywords_in_excel(keywords, folder_path):
    # 遍历文件夹中的所有文件和子文件夹
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.xlsx'):
                # print(f"正在处理文件：{file}")
                file_path = os.path.join(root, file)
                # 打开 Excel 文件
                workbook = load_workbook(filename=file_path)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    # 遍历每个单元格，查找关键词
                    for row in sheet.iter_rows():
                        row_data = [cell.value for cell in row if cell.value is not None]
                        if row_data and any(keyword in str(cell.value) for cell in row for keyword in keywords):
                            print(f"关键词 '{keywords}' 在文件 '{file}'的数据为: {row_data}")


if __name__ == '__main__':
    # 设置要搜索的关键词列表和包含 Excel 文件的文件夹路径
    keywords_to_search = input("请输入要搜索的关键词，多个关键词用空格分隔：")
    keywords_to_search = keywords_to_search.split()
    folder_to_search = "path"
    # 调用函数进行搜索
    search_keywords_in_excel(keywords_to_search, folder_to_search)
